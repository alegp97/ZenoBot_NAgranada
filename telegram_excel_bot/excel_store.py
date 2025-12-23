import os
from typing import Any, Optional

from filelock import FileLock
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# Cabeceras canónicas (humanas)
HEADERS = [
  "id", "Título", "Autor", "Procedencia", "Categoría",
  "Editorial", "Año", "Columna", "Fila", "ISBN",
  "F_revision", "Comentarios"
]


# Mapa de normalización: Excel → clave canónica
HEADER_MAP = {
    "id": "id",

    "título": "Título",
    "titulo": "Título",

    "autor": "Autor",
    "editorial": "Editorial",

    "año": "Año",
    "ano": "Año",

    "columna": "Columna",
    "fila": "Fila",

    "isbn": "ISBN",

    "procedencia": "Procedencia",
    "categoría": "Categoría",
    "categoria": "Categoría",
    "comentarios": "Comentarios",
    "comentario": "Comentarios",

    "F.revision": "F_revision",
    "f.revisión": "F_revision",
    "f_revision": "F_revision",
    "F_Revisión": "F_revision",
    "F.Revisión": "F_revision",
    "f revisión": "F_revision",
    "fecha revision": "F_revision",
    "fecha revisión": "F_revision",
    "frevision": "F_revision",
    "f_revisión": "F_revision",
}



class ExcelStore:
    def __init__(self, path: str, sheet: str):
        self.path = path
        self.sheet = sheet
        self.lock_path = path + ".lock"

        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        if not os.path.exists(path):
            self._init_book()

    # ---------- inicialización ----------

    def _init_book(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet
        ws.append(HEADERS)
        wb.save(self.path)

    def _open(self) -> tuple[Any, Worksheet]:
        wb = load_workbook(self.path)
        if self.sheet not in wb.sheetnames:
            ws = wb.create_sheet(self.sheet)
            ws.append(HEADERS)
        else:
            ws = wb[self.sheet]
            if ws.max_row < 1:
                ws.append(HEADERS)

        idx = self._header_index(ws)
        self._validate_headers(idx)
        return wb, ws

    # ---------- cabeceras ----------

    def _header_index(self, ws: Worksheet) -> dict[str, int]:
        idx: dict[str, int] = {}

        for i, cell in enumerate(ws[1], start=1):
            if cell.value is None:
                continue

            raw = str(cell.value).strip().lower()
            key = HEADER_MAP.get(raw)
            if key:
                idx[key] = i

        return idx

    def _validate_headers(self, idx: dict[str, int]) -> None:
        required = set(HEADERS)
        missing = required - set(idx.keys())
        if missing:
            raise RuntimeError(f"Faltan columnas obligatorias en Excel: {missing}")



    # ---------- utilidades ----------
    
    def _next_id(self, ws: Worksheet, idx: dict[str, int]) -> str:
        col_id = idx["id"]
        last = 0

        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, col_id).value
            if not v:
                continue
            s = str(v).strip()
            if s.startswith("B") and s[1:].isdigit():
                last = max(last, int(s[1:]))

        return f"B{last + 1:06d}"

    def _row_to_dict(self, ws: Worksheet, r: int, idx: dict[str, int]) -> dict[str, Any]:
        return {h: ws.cell(r, idx[h]).value for h in HEADERS}



    # ---------- operaciones públicas ----------

    def add(self, book: dict[str, Any]) -> str:
        """
        Append puro:
        - Inserta siempre al final (push)
        - id = (fila_excel - 1) porque fila 1 es cabecera
        book keys (internos): titulo, autor, editorial, ano, columna, fila, isbn
        """
        with FileLock(self.lock_path):
            wb, ws = self._open()
            idx = self._header_index(ws)

            # siguiente fila real donde se va a escribir (append)
            excel_row = ws.max_row + 1
            new_id = int(excel_row - 1)

            row = [""] * len(HEADERS)
            row[idx["id"] - 1] = new_id
            row[idx["Título"] - 1] = book.get("titulo", "") or ""
            row[idx["Autor"] - 1] = book.get("autor", "") or ""
            row[idx["Editorial"] - 1] = book.get("editorial", "") or ""
            row[idx["Año"] - 1] = book.get("ano", None)
            row[idx["Columna"] - 1] = book.get("columna", None)
            row[idx["Fila"] - 1] = book.get("fila", None)
            row[idx["ISBN"] - 1] = book.get("isbn", "") or ""

            ws.append(row)
            wb.save(self.path)
            return new_id


    def get_by_id(self, book_id: str) -> Optional[dict[str, Any]]:
        if not book_id:
            return None

        with FileLock(self.lock_path):
            wb, ws = self._open()
            idx = self._header_index(ws)
            col_id = idx["id"]

            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, col_id).value
                if v and str(v).strip() == str(book_id).strip():
                    return self._row_to_dict(ws, r, idx)

            return None

    def find(self, criteria: dict[str, str], limit: int = 20) -> list[dict[str, Any]]:
        limit = max(1, min(int(limit), 50))
        crit = {k: v.strip().lower() for k, v in criteria.items() if v and v.strip()}
        if not crit:
            return []

        key_to_header = {
            "titulo": "Título",
            "autor": "Autor",
            "editorial": "Editorial",
            "ano": "Año",
            "isbn": "ISBN",
            "fila": "Fila",
            "columna": "Columna",
            "id": "id",
        }

        with FileLock(self.lock_path):
            wb, ws = self._open()
            idx = self._header_index(ws)

            out: list[dict[str, Any]] = []
            for r in range(2, ws.max_row + 1):
                rowd = self._row_to_dict(ws, r, idx)
                ok = True

                for k, needle in crit.items():
                    h = key_to_header.get(k)
                    if not h:
                        ok = False
                        break
                    hay = "" if rowd.get(h) is None else str(rowd.get(h)).lower()
                    if needle not in hay:
                        ok = False
                        break

                if ok:
                    out.append(rowd)
                    if len(out) >= limit:
                        break

            return out

    def last(self, n: int = 10) -> list[dict[str, Any]]:
        n = max(1, min(int(n), 200))

        with FileLock(self.lock_path):
            wb, ws = self._open()
            idx = self._header_index(ws)

            start = max(2, ws.max_row - n + 1)
            return [self._row_to_dict(ws, r, idx) for r in range(start, ws.max_row + 1)]

        
    def update_fields(self, book_id: str, changes: dict[str, Any]) -> bool:
        """
        changes usa keys internas: titulo, autor, editorial, ano, fila, columna, isbn
        - strings: "" para vaciar
        - ints: null para vaciar
        """
        key_to_headers = {
            "titulo": ["Título", "Titulo"],
            "autor": ["Autor"],
            "procedencia": ["Procedencia"],
            "categoria": ["Categoría", "Categoria"],
            "editorial": ["Editorial"],
            "ano": ["Año", "Ano"],
            "columna": ["Columna"],
            "fila": ["Fila"],
            "isbn": ["ISBN"],
            "f_revision": ["F_revisión", "F_revision", "F_revision _", "F_revision->"],
            "comentarios": ["Comentarios"],
        }


        with FileLock(self.lock_path):
            wb, ws = self._open()
            idx = self._header_index(ws)

            col_id = idx["id"]

            # localizar fila por id
            target_row = None
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, col_id).value
                if v and str(v).strip() == str(book_id).strip():
                    target_row = r
                    break

            if target_row is None:
                return False

            # aplicar cambios
            for k, v in changes.items():
                if k not in key_to_headers:
                    continue

                # elige la primera cabecera que exista en el Excel
                header = next((h for h in key_to_headers[k] if h in idx), None)
                if not header:
                    continue

                c = idx[header]

                if k in ("fila", "columna", "ano"):
                    ws.cell(target_row, c).value = None if v is None else int(v)
                else:
                    ws.cell(target_row, c).value = "" if v is None else str(v)


            wb.save(self.path)
            return True

    def delete_and_compact(self, book_id: int) -> bool:
        """
        Borra la fila del libro con id=book_id y luego recalcula todos los ids para que:
        id = (fila_excel - 1)
        """
        def _same_id(cell_value: Any, book_id: Any) -> bool:
            try:
                return int(cell_value) == int(book_id)
            except (TypeError, ValueError):
                return False

        with FileLock(self.lock_path):
            wb, ws = self._open()
            idx = self._header_index(ws)

            col_id = idx["id"]

            # 1) localizar fila a borrar
            delete_row = None
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, col_id).value
                if _same_id(v, book_id):
                    delete_row = r
                    break

            if delete_row is None:
                return False

            # 2) borrar fila (desplaza hacia arriba)
            ws.delete_rows(delete_row, 1)

            # 3) compactar ids: id = fila - 1
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col_id).value = r - 1

            wb.save(self.path)
            return True
