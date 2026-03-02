"""
Tests for telegram_excel_bot.excel_store — ExcelStore CRUD + filtering.
Uses a temporary Excel file (no real catalogo.xlsx needed).
"""
import pytest


# ═══════════════════════════════════════════════════════════════
# Initialization
# ═══════════════════════════════════════════════════════════════


class TestInit:
    def test_creates_file(self, store, tmp_excel):
        """ExcelStore creates the file if it doesn't exist."""
        import os
        path, _ = tmp_excel
        assert os.path.exists(path)

    def test_headers_present(self, store):
        """The first row must contain all canonical headers."""
        from telegram_excel_bot.excel_store import HEADERS
        from openpyxl import load_workbook

        wb = load_workbook(store.path)
        ws = wb[store.sheet]
        row1 = [c.value for c in ws[1]]
        for h in HEADERS:
            assert h in row1, f"Missing header: {h}"


# ═══════════════════════════════════════════════════════════════
# Add
# ═══════════════════════════════════════════════════════════════


class TestAdd:
    def test_add_returns_id(self, store, sample_book):
        new_id = store.add(sample_book)
        assert new_id is not None

    def test_add_increments_id(self, store, sample_book):
        id1 = store.add(sample_book)
        id2 = store.add(sample_book)
        assert int(id2) == int(id1) + 1

    def test_add_saves_fields(self, store, sample_book):
        new_id = store.add(sample_book)
        row = store.get_by_id(new_id)
        assert row is not None
        assert row["Título"] == "República"
        assert row["Autor"] == "Platón"
        assert row["Editorial"] == "Gredos"

    def test_add_minimal_book(self, store):
        """Only titulo is required."""
        new_id = store.add({"titulo": "Sin Autor"})
        row = store.get_by_id(new_id)
        assert row["Título"] == "Sin Autor"
        assert row["Autor"] in (None, "")


# ═══════════════════════════════════════════════════════════════
# Get by ID
# ═══════════════════════════════════════════════════════════════


class TestGetById:
    def test_get_existing(self, store, sample_book):
        new_id = store.add(sample_book)
        row = store.get_by_id(new_id)
        assert row is not None
        assert row["Título"] == "República"

    def test_get_nonexistent(self, store):
        assert store.get_by_id(9999) is None

    def test_get_none(self, store):
        assert store.get_by_id(None) is None

    def test_get_empty(self, store):
        assert store.get_by_id("") is None


# ═══════════════════════════════════════════════════════════════
# Find (original fields)
# ═══════════════════════════════════════════════════════════════


class TestFind:
    def test_find_by_titulo(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"titulo": "República"})
        assert len(res) == 1
        assert res[0]["Título"] == "República"

    def test_find_by_autor(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"autor": "Platón"})
        assert len(res) == 2  # República + Fedón

    def test_find_by_editorial(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"editorial": "Gredos"})
        assert len(res) == 3  # República, Ética, Enquiridión

    def test_find_by_isbn(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"isbn": "978-84-249-1027-3"})
        assert len(res) == 1

    def test_find_by_ano(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"ano": "350"})
        assert len(res) == 1
        assert res[0]["Autor"] == "Aristóteles"

    def test_find_no_results(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"titulo": "No Existe"})
        assert len(res) == 0

    def test_find_empty_criteria(self, store_with_books):
        store, _ = store_with_books
        res = store.find({})
        assert len(res) == 0

    def test_find_limit(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"editorial": "Gredos"}, limit=2)
        assert len(res) == 2

    def test_find_multiple_criteria(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"autor": "Platón", "editorial": "Alianza"})
        assert len(res) == 1
        assert res[0]["Título"] == "Fedón"

    def test_find_partial_match(self, store_with_books):
        """find uses 'contains' matching (substring)."""
        store, _ = store_with_books
        res = store.find({"titulo": "edi"})  # should match "Meditaciones"
        assert len(res) == 1
        assert "Meditaciones" in res[0]["Título"]


# ═══════════════════════════════════════════════════════════════
# Find (NEW fields: procedencia, categoria, f_revision, comentarios)
# ═══════════════════════════════════════════════════════════════


class TestFindNewFields:
    def test_find_by_procedencia(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"procedencia": "Donación"})
        assert len(res) == 3  # República, Meditaciones, Fedón

    def test_find_by_categoria(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"categoria": "Filosofía"})
        assert len(res) == 3  # República, Ética, Fedón

    def test_find_by_categoria_estoicismo(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"categoria": "Estoicismo"})
        assert len(res) == 2  # Meditaciones, Enquiridión

    def test_find_by_comentarios(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"comentarios": "manuscrito"})
        assert len(res) == 1
        assert res[0]["Autor"] == "Marco Aurelio"

    def test_find_by_f_revision(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"f_revision": "2025"})
        assert len(res) == 1
        assert res[0]["Autor"] == "Aristóteles"

    def test_find_combined_new_and_old(self, store_with_books):
        """Combine a new field (procedencia) with an old field (editorial)."""
        store, _ = store_with_books
        res = store.find({"procedencia": "Donación", "editorial": "Alianza"})
        assert len(res) == 2  # Meditaciones, Fedón

    def test_find_by_fila(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"fila": "5"})
        assert len(res) == 1
        assert res[0]["Título"] == "República"

    def test_find_by_columna(self, store_with_books):
        store, _ = store_with_books
        res = store.find({"columna": "3"})
        assert len(res) == 2  # Meditaciones, Enquiridión


# ═══════════════════════════════════════════════════════════════
# Update
# ═══════════════════════════════════════════════════════════════


class TestUpdate:
    def test_update_titulo(self, store, sample_book):
        new_id = store.add(sample_book)
        ok = store.update_fields(new_id, {"titulo": "La República"})
        assert ok
        row = store.get_by_id(new_id)
        assert row["Título"] == "La República"

    def test_update_multiple_fields(self, store, sample_book):
        new_id = store.add(sample_book)
        ok = store.update_fields(new_id, {"autor": "Plato", "editorial": "Penguin"})
        assert ok
        row = store.get_by_id(new_id)
        assert row["Autor"] == "Plato"
        assert row["Editorial"] == "Penguin"

    def test_update_procedencia(self, store, sample_book):
        new_id = store.add(sample_book)
        ok = store.update_fields(new_id, {"procedencia": "Compra"})
        assert ok
        row = store.get_by_id(new_id)
        assert row["Procedencia"] == "Compra"

    def test_update_categoria(self, store, sample_book):
        new_id = store.add(sample_book)
        ok = store.update_fields(new_id, {"categoria": "Metafísica"})
        assert ok
        row = store.get_by_id(new_id)
        assert row["Categoría"] == "Metafísica"

    def test_update_f_revision(self, store, sample_book):
        new_id = store.add(sample_book)
        ok = store.update_fields(new_id, {"f_revision": "02/03/2026"})
        assert ok
        row = store.get_by_id(new_id)
        assert row["F_revision"] == "02/03/2026"

    def test_update_comentarios(self, store, sample_book):
        new_id = store.add(sample_book)
        ok = store.update_fields(new_id, {"comentarios": "Nueva nota"})
        assert ok
        row = store.get_by_id(new_id)
        assert row["Comentarios"] == "Nueva nota"

    def test_update_nonexistent_returns_false(self, store):
        ok = store.update_fields(9999, {"titulo": "X"})
        assert not ok

    def test_update_numeric_fields(self, store, sample_book):
        new_id = store.add(sample_book)
        ok = store.update_fields(new_id, {"fila": 10, "columna": 20, "ano": 2026})
        assert ok
        row = store.get_by_id(new_id)
        assert row["Fila"] == 10
        assert row["Columna"] == 20
        assert row["Año"] == 2026


# ═══════════════════════════════════════════════════════════════
# Delete & Compact
# ═══════════════════════════════════════════════════════════════


class TestDelete:
    def test_delete_existing(self, store, sample_book):
        new_id = store.add(sample_book)
        ok = store.delete_and_compact(new_id)
        assert ok
        assert store.get_by_id(new_id) is None

    def test_delete_nonexistent(self, store):
        ok = store.delete_and_compact(9999)
        assert not ok

    def test_delete_compacts_ids(self, store, sample_book):
        id1 = store.add(sample_book)
        id2 = store.add({**sample_book, "titulo": "Libro 2"})
        id3 = store.add({**sample_book, "titulo": "Libro 3"})

        store.delete_and_compact(id2)

        # After compact, id3 should now have id = old_id3 - 1
        row = store.get_by_id(int(id3) - 1)
        assert row is not None
        assert row["Título"] == "Libro 3"


# ═══════════════════════════════════════════════════════════════
# Last
# ═══════════════════════════════════════════════════════════════


class TestLast:
    def test_last_returns_n(self, store_with_books):
        store, _ = store_with_books
        res = store.last(3)
        assert len(res) == 3

    def test_last_ordering(self, store_with_books):
        store, ids = store_with_books
        res = store.last(2)
        # Last 2 should include the last two books added
        titles = {r["Título"] for r in res}
        assert "Enquiridión" in titles
        assert "Fedón" in titles

    def test_last_exceeding(self, store_with_books):
        store, _ = store_with_books
        res = store.last(100)
        assert len(res) == 5  # only 5 books exist


# ═══════════════════════════════════════════════════════════════
# Get All
# ═══════════════════════════════════════════════════════════════


class TestGetAll:
    def test_get_all_empty(self, store):
        res = store.get_all()
        assert res == []

    def test_get_all_with_books(self, store_with_books):
        store, _ = store_with_books
        res = store.get_all()
        assert len(res) == 5


# ═══════════════════════════════════════════════════════════════
# Search by Field (advanced ops)
# ═══════════════════════════════════════════════════════════════


class TestSearchByField:
    def test_eq(self, store_with_books):
        store, _ = store_with_books
        res = store.search_by_field("Autor", "eq", "Platón")
        assert len(res) == 2

    def test_contains(self, store_with_books):
        store, _ = store_with_books
        res = store.search_by_field("Título", "contains", "edi")
        assert len(res) == 1  # Meditaciones

    def test_startswith(self, store_with_books):
        store, _ = store_with_books
        res = store.search_by_field("Título", "startswith", "Re")
        assert len(res) == 1

    def test_endswith(self, store_with_books):
        store, _ = store_with_books
        res = store.search_by_field("Título", "endswith", "ón")
        assert len(res) == 2  # Enquiridión, Fedón

    def test_neq(self, store_with_books):
        store, _ = store_with_books
        res = store.search_by_field("Editorial", "neq", "Gredos")
        assert len(res) == 2  # Alianza books

    def test_no_match(self, store_with_books):
        store, _ = store_with_books
        res = store.search_by_field("Autor", "eq", "Descartes")
        assert len(res) == 0



